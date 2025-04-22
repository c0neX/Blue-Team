#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Script to map IOCs (Indicators of Compromise) to MITRE ATT&CK TTPs (Tactics, Techniques, and Procedures)
using data from the APTnotes repository.

Author: c0neX
Version: 2.0
"""

import pandas as pd
import requests
import re
import matplotlib.pyplot as plt
import seaborn as sns
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo
from io import StringIO
import sys
from tqdm import tqdm
import argparse
import os
import time
from datetime import datetime

# ------------------- Configuration -------------------
# Note: This URL was updated based on previous discussion.
# Ensure it still points to the correct raw CSV data if the repository structure changes again.
APTNOTES_URL = "https://raw.githubusercontent.com/aptnotes/data/refs/heads/master/APTnotes.csv"
TTP_REGEX = re.compile(r'\bT\d{4}(?:\.\d{3})?\b')
DEFAULT_OUTPUT_EXCEL = "ioc_ttp_mapping.xlsx"
DEFAULT_OUTPUT_HEATMAP = "ttp_heatmap.png"
CACHE_FILE = "aptnotes_cache.csv"
CACHE_TIME = 24  # hours

# ------------------- Helper Functions -------------------

def normalize_ioc(ioc):
    """Normalizes IOCs for better matching"""
    ioc = str(ioc).lower().strip()
    # Remove protocols and paths from URLs
    if ioc.startswith(('http://', 'https://')):
        ioc = ioc.split('//')[1].split('/')[0]
    # Remove special characters (be careful with what characters are allowed in real IOCs)
    # This regex is a basic example and might need tuning based on expected IOC types.
    return re.sub(r'[^a-z0-9.:/-]', '', ioc)

def enhance_ttp_extraction(text):
    """Enhances TTP extraction"""
    text = str(text).upper()
    # Additional patterns to capture TTPs
    patterns = [
        r'\bT\d{4}(?:\.\d{3})?\b',  # TXXXX or TXXXX.XXX
        r'ATT&CK T\d{4}',           # ATT&CK TXXXX
        r'TECHNIQUE T\d{4}'         # TECHNIQUE TXXXX (ensure case-insensitive search or handle cases)
    ]
    ttps = set()
    for pattern in patterns:
        # Using re.IGNORECASE might be better if not upper-casing the text
        ttps.update(re.findall(pattern, text))
    # Cleaning results
    return [re.sub(r'[^T0-9.]', '', ttp) for ttp in ttps if ttp.startswith('T')]

def load_iocs_from_csv(filepath):
    """Loads IOCs from CSV file with validation and auto-detection"""
    try:
        df = pd.read_csv(filepath)

        # Attempt to auto-detect IOC column
        ioc_col = None
        for col in df.columns:
            if 'ioc' in col.lower() or 'indicator' in col.lower() or 'value' in col.lower(): # Added 'value' as a potential column
                ioc_col = col
                break

        if ioc_col is None:
            # Fallback to first column if detection fails
            if len(df.columns) > 0:
                ioc_col = df.columns[0]
                print(f"[!] IOC column not found, using first column: '{ioc_col}'")
            else:
                 raise ValueError("CSV file is empty or has no columns.")

        iocs = df[ioc_col].dropna().astype(str).str.strip().unique().tolist()
        print(f"[+] {len(iocs)} unique IOCs loaded")
        if len(iocs) > 10:
            print(f"[*] Example IOCs: {iocs[:5]} [...]")
        else:
            print(f"[*] Loaded IOCs: {iocs}")

        return iocs

    except FileNotFoundError:
         raise FileNotFoundError(f"Error: Input CSV file not found at '{filepath}'")
    except pd.errors.EmptyDataError:
         raise ValueError(f"Error: Input CSV file '{filepath}' is empty.")
    except Exception as e:
        raise ValueError(f"Error loading IOCs: {str(e)}")

def fetch_aptnotes_data(url=APTNOTES_URL, cache_time=CACHE_TIME):
    """Fetches APTnotes data with caching support"""
    use_cache = False

    if os.path.exists(CACHE_FILE):
        mod_time = os.path.getmtime(CACHE_FILE)
        if (time.time() - mod_time) < cache_time * 3600:
            use_cache = True

    if use_cache:
        print("[*] Using cached APTnotes data")
        try:
            return pd.read_csv(CACHE_FILE)
        except Exception as e:
            print(f"[!] Error reading cache file, attempting download: {str(e)}")
            # Fall through to download if cache read fails

    print("[*] Downloading updated APTnotes data...")
    try:
        resp = requests.get(url, timeout=30)
        resp.raise_for_status() # Raise an HTTPError for bad responses (4xx or 5xx)
        df = pd.read_csv(StringIO(resp.text))

        # Save cache
        df.to_csv(CACHE_FILE, index=False)
        return df

    except requests.exceptions.RequestException as e:
        if os.path.exists(CACHE_FILE):
            print(f"[!] Network error ({e}), using cache as fallback.")
            try:
                return pd.read_csv(CACHE_FILE)
            except Exception as cache_e:
                 raise ConnectionError(f"Network error AND failed to read cache: {str(e)} | Cache error: {str(cache_e)}")
        else:
             raise ConnectionError(f"Network error fetching APTnotes data and no cache available: {str(e)}")
    except Exception as e:
        raise RuntimeError(f"Error processing downloaded APTnotes data: {str(e)}")


def search_iocs_and_ttps(df, iocs):
    """Searches for IOCs and extracts associated TTPs"""
    # --- Adjusted column names based on the CSV structure provided by the user ---
    # We now use 'Title' and 'Filename' as the content sources, as the full text
    # is likely not in the CSV. The ability to find IOCs and TTPs will be limited
    # to these fields.
    required_columns = ["Title", "Link", "Filename", "Date"]
    for col in required_columns:
        if col not in df.columns:
            raise ValueError(f"Required column missing in APTnotes data: '{col}'. Available columns: {list(df.columns)}")

    # Using .copy() to avoid SettingWithCopyWarning
    df = df.dropna(subset=[col for col in required_columns if col != "Link"]).copy() # Link might be empty
    
    # Combine relevant text fields for searching
    # Note: This is a limitation. Searching only Title/Filename might miss many IOCs/TTPs.
    # A more complex solution would involve downloading/parsing linked documents.
    df['SearchContent'] = df['Title'].fillna('') + ' ' + df['Filename'].fillna('')

    normalized_iocs = [(ioc, normalize_ioc(ioc)) for ioc in iocs if ioc and str(ioc).strip()]

    matches = []
    unmatched_iocs = set(iocs) # Track which IOCs were not matched
    print("[*] Searching for matches with IOCs...")

    # Use tqdm for progress bar
    for _, row in tqdm(df.iterrows(), total=len(df), desc="Analyzing reports"):
        content = str(row['SearchContent']).lower()
        norm_content = normalize_ioc(content) # Normalize combined content

        # Extract TTPs from the available content (Title + Filename)
        # This assumes TTPs are mentioned in these fields, which might not always be true.
        report_ttps = enhance_ttp_extraction(content)

        if not report_ttps:
            continue # Skip reports without any extracted TTPs in these fields

        for original_ioc, norm_ioc in normalized_iocs:
            # --- Multiple matching strategies ---
            # Check if the original or normalized IOC appears in the content
            # Added checks for domains/subdomains and partial hash/string matches
            is_match = False
            if original_ioc.lower() in content:
                 is_match = True
            elif norm_ioc and len(norm_ioc) > 2 and norm_ioc in norm_content: # Check normalized content if normalization resulted in changes
                 is_match = True
            # Add more sophisticated checks here if needed, e.g., IP regex match, domain suffix match

            if is_match:
                unmatched_iocs.discard(original_ioc) # Remove matched IOC from the unmatched set
                for ttp in report_ttps:
                    matches.append({
                        "IOC": original_ioc,
                        "TTP": ttp,
                        "Technique": f"T{ttp[1:5]}" if len(ttp) >= 5 and ttp[1:5].isdigit() else ttp, # Group by main technique TXXXX
                        "Subtechnique": ttp if '.' in ttp else '',
                        "Title": row.get("Title", ""),
                        "Reference": row.get("Link", ""), # Using 'Link' as reference
                        "File": row.get("Filename", ""), # Using 'Filename'
                        "Date": row.get("Date", "")
                    })

    # Create DataFrame for unmatched IOCs
    unmatched_df = pd.DataFrame({
        "IOC": list(unmatched_iocs),
        "Reason": "Not found in analyzed report metadata (Title/Filename)"
    }) if unmatched_iocs else pd.DataFrame()


    if not matches:
        print("[!] No TTPs mapped from provided IOCs based on available report metadata.")
        return pd.DataFrame(), unmatched_df

    result_df = pd.DataFrame(matches).drop_duplicates()

    # Add links to MITRE ATT&CK
    base_url = "https://attack.mitre.org/techniques/"
    result_df["ATT&CK Link"] = result_df["TTP"].apply(
        lambda x: f"{base_url}{x.replace('.', '/')}")

    return result_df, unmatched_df


def generate_heatmap(mapping_df, out_path=DEFAULT_OUTPUT_HEATMAP):
    """Generates heatmap visualization"""
    if mapping_df.empty:
        print("[!] No data to generate heatmap")
        return

    # Group by main technique for heatmap to avoid too many columns from sub-techniques
    if 'Technique' not in mapping_df.columns:
         print("[!] 'Technique' column missing, cannot generate heatmap.")
         return

    pivot = pd.pivot_table(mapping_df, index="IOC", columns="Technique",
                           aggfunc='size', fill_value=0)

    if pivot.empty:
        print("[!] Invalid data for heatmap generation")
        return

    # Order by total occurrences across techniques
    pivot['Total'] = pivot.sum(axis=1)
    pivot = pivot.sort_values(by='Total', ascending=False).drop(columns='Total')
    # Order columns alphabetically
    pivot = pivot[pivot.columns.sort_values()]

    plt.figure(figsize=(max(12, len(pivot.columns)*0.8), max(5, len(pivot)*0.4))) # Adjust size based on data
    sns.heatmap(pivot, cmap="YlOrRd", annot=True, fmt="d",
                cbar_kws={"label": "Occurrences"},
                linewidths=.5)
    plt.title(f"ATT&CK Technique Heatmap by IOC ({datetime.now().strftime('%Y-%m-%d')})")
    plt.ylabel("IOC")
    plt.xlabel("ATT&CK Technique ID")
    plt.tight_layout()

    try:
        plt.savefig(out_path, dpi=300)
        print(f"[+] Heatmap saved to {out_path}")
    except Exception as e:
        print(f"[!] Error saving heatmap: {str(e)}")
    finally:
        plt.close()

def export_to_excel(mapped_df, unmatched_df=None, filename=DEFAULT_OUTPUT_EXCEL):
    """Exports results to Excel"""
    if mapped_df.empty and (unmatched_df is None or unmatched_df.empty):
        print("[!] No data to export")
        return

    try:
        wb = Workbook()
        ws_count = 0 # Track number of sheets added

        # Main Mappings Sheet
        if not mapped_df.empty:
            ws = wb.active # Get the default sheet
            ws.title = "Mappings"
            ws_count += 1

            # Sort by IOC and TTP for better readability
            mapped_df = mapped_df.sort_values(by=["IOC", "TTP"])

            for r_idx, row in enumerate(dataframe_to_rows(mapped_df, index=False, header=True)):
                 ws.append(row)
            
            # Apply table formatting if data exists beyond header
            if len(mapped_df) > 0:
                tab = Table(displayName="MappingsTable", ref=f"A1:{get_column_letter(ws.max_column)}{ws.max_row}")
                style = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
                tab.tableStyleInfo = style
                ws.add_table(tab)

                # Auto-fit column widths
                for col in ws.columns:
                    max_length = 0
                    column = col[0].column_letter
                    for cell in col:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(cell.value)
                        except:
                            pass
                    adjusted_width = (max_length + 2)
                    # Limit max width to prevent excessively wide columns
                    ws.column_dimensions[column].width = min(adjusted_width, 75) # Limit to 75


        # Unmatched IOCs Sheet
        if unmatched_df is not None and not unmatched_df.empty:
            if ws_count == 0: # If no mappings, use the default sheet for unmatched
                ws_unmatched = wb.active
                ws_unmatched.title = "Unmatched IOCs"
            else: # Otherwise, create a new sheet
                ws_unmatched = wb.create_sheet("Unmatched IOCs")
            ws_count += 1

            for row in dataframe_to_rows(unmatched_df, index=False, header=True):
                ws_unmatched.append(row)
            
            if len(unmatched_df) > 0:
                tab_unmatched = Table(displayName="UnmatchedIOCsTable", ref=f"A1:{get_column_letter(ws_unmatched.max_column)}{ws_unmatched.max_row}")
                tab_unmatched.tableStyleInfo = style
                ws_unmatched.add_table(tab_unmatched)

                for col in ws_unmatched.columns:
                     max_length = 0
                     column = col[0].column_letter
                     for cell in col:
                         try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(cell.value)
                         except:
                             pass
                     adjusted_width = (max_length + 2)
                     ws_unmatched.column_dimensions[column].width = min(adjusted_width, 75) # Limit to 75


        # Metadata Sheet
        ws_meta = wb.create_sheet("Metadata")
        ws_count += 1
        ws_meta.append(["Generated On", datetime.now().strftime('%Y-%m-%d %H:%M:%S')])
        ws_meta.append(["Total IOCs Processed", len(iocs) if 'iocs' in locals() else 0]) # Check if iocs variable exists
        ws_meta.append(["Total Mappings Found", len(mapped_df) if not mapped_df.empty else 0])
        ws_meta.append(["Unique TTPs Identified", len(mapped_df['TTP'].unique()) if not mapped_df.empty else 0])
        ws_meta.append(["Reports Analyzed", len(mapped_df['File'].unique()) if not mapped_df.empty and 'File' in mapped_df.columns else 0])
        ws_meta.append(["APTnotes Data Source", APTNOTES_URL])
        ws_meta.append(["IOC File Source", args.ioc_file if 'args' in locals() else 'N/A']) # Check if args variable exists

        # Remove default blank sheet if other sheets were added
        if ws_count > 1 and "Sheet" in wb.sheetnames and wb["Sheet"].max_row == 0 and wb["Sheet"].max_column == 0:
             wb.remove(wb["Sheet"])

        wb.save(filename)
        print(f"[+] Excel report saved to {filename}")

    except NameError:
         print(f"[!] Error during Excel export: Variables 'iocs' or 'args' not defined. This might happen if the script failed early.")
    except Exception as e:
        print(f"[!] Error exporting to Excel: {str(e)}")

# Helper for Excel column letters
from openpyxl.utils import get_column_letter


def parse_arguments():
    """Parses command line arguments"""
    parser = argparse.ArgumentParser(
        description="Maps IOCs to ATT&CK TTPs using APTnotes data",
        formatter_class=argparse.ArgumentDefaultsHelpFormatter)
    parser.add_argument("ioc_file", help="CSV file containing IOCs")
    parser.add_argument("-o", "--output", help="Base name for output files (e.g., 'results' will produce results.xlsx and results.png)",
                        default="ioc_ttp_results")
    parser.add_argument("--no-heatmap", action="store_true",
                        help="Do not generate visual heatmap")
    parser.add_argument("--debug", action="store_true",
                        help="Enable debug mode with detailed information")
    parser.add_argument("--cache-time", type=int, default=CACHE_TIME,
                        help="APTnotes data cache time in hours (0 to disable cache)")
    return parser.parse_args()

# ------------------- Main -------------------

def main():
    args = parse_arguments()

    # Access cache_time from args
    global CACHE_TIME
    CACHE_TIME = args.cache_time

    try:
        print(f"\n{'='*50}")
        print(f" IOC-TTP Mapper v2.0 - {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        print(f"{'='*50}\n")

        # Load IOCs
        print(f"[*] Loading IOCs from {args.ioc_file}...")
        # Pass args to load_iocs_from_csv if needed, though it currently only needs filepath
        iocs = load_iocs_from_csv(args.ioc_file)
        if not iocs:
             print("[!] No valid IOCs found in the input file. Exiting.")
             # Still attempt to export an empty report or a report with unmatched IOCs if load succeeded but found none
             export_to_excel(pd.DataFrame(), pd.DataFrame({"IOC": [], "Reason": []}), f"{args.output}.xlsx")
             return


        # Fetch APTnotes data
        print("\n[*] Fetching APTnotes data...")
        # Pass cache_time from args
        apt_df = fetch_aptnotes_data(cache_time=args.cache_time)
        print(f"[+] {len(apt_df)} reports loaded")

        if args.debug:
            print("\n[DEBUG] Example reports:")
            # Adjusted column names for debug output
            debug_cols = [col for col in ['Title', 'Date', 'Source', 'Link'] if col in apt_df.columns]
            if debug_cols:
                print(apt_df[debug_cols].head().to_string())
            else:
                 print("[DEBUG] No common columns ('Title', 'Date', 'Source', 'Link') found in APTnotes data.")


        # Map IOCs to TTPs
        print("\n[*] Mapping IOCs to TTPs...")
        # Pass iocs from loaded data
        mapped_df, unmatched_df = search_iocs_and_ttps(apt_df, iocs)


        # Generate outputs
        excel_file = f"{args.output}.xlsx"
        heatmap_file = f"{args.output}.png"

        # Export to Excel (always attempt to export, even if no mappings, to show unmatched)
        print(f"\n[*] Exporting results to {excel_file}...")
        # Pass loaded iocs and args to export_to_excel for metadata sheet
        export_to_excel(mapped_df, unmatched_df, excel_file)

        if not mapped_df.empty:
            print(f"\n[+] {len(mapped_df)} mappings found")
            if 'TTP' in mapped_df.columns:
                print(f"[+] {len(mapped_df['TTP'].unique())} unique TTPs identified")

            # Generate heatmap (only if mappings were found and --no-heatmap is not set)
            if not args.no_heatmap:
                print(f"[*] Generating heatmap to {heatmap_file}...")
                generate_heatmap(mapped_df, heatmap_file)
            else:
                print("[*] Heatmap generation skipped (--no-heatmap flag set).")
        else:
             print("\n[!] No mappings found between provided IOCs and APTnotes data.")
             print(f"[*] Check '{excel_file}' for a list of unmatched IOCs.")


        print("\n[✓] Analysis complete!")

    except FileNotFoundError as e:
         print(f"\n[!] ERROR: Input file not found - {str(e)}")
         sys.exit(1)
    except ValueError as e:
        print(f"\n[!] ERROR: Data format issue - {str(e)}")
        sys.exit(1)
    except ConnectionError as e:
        print(f"\n[!] ERROR: Network or cache issue - {str(e)}")
        sys.exit(1)
    except Exception as e:
        # Catch any other unexpected errors
        print(f"\n[!] An unexpected ERROR occurred: {str(e)}")
        if args.debug:
            import traceback
            traceback.print_exc() # Print full traceback in debug mode
        sys.exit(1)


if __name__ == "__main__":
    main()
